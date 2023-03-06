from openpyxl import load_workbook
from matplotlib import pyplot as plt

# --- UNCOMMENT TO: Plot data in graph ---
# wb1 = load_workbook(filename='../Training_Results/1. Optimizer/ada_results.xlsx')
# wb2 = load_workbook(filename='../Training_Results/1. Optimizer/adam_results.xlsx')
# wb3 = load_workbook(filename='../Training_Results/1. Optimizer/rms_results.xlsx')

# wb1 = load_workbook(filename='../Training_Results/2. Learning Rate/0.01.xlsx')
# wb2 = load_workbook(filename='../Training_Results/2. Learning Rate/0.001.xlsx')
# wb3 = load_workbook(filename='../Training_Results/2. Learning Rate/0.0001.xlsx')

# wb1 = load_workbook(filename='../Training_Results/3. Batch Size/32.xlsx')
# wb2 = load_workbook(filename='../Training_Results/3. Batch Size/64.xlsx')
# wb3 = load_workbook(filename='../Training_Results/3. Batch Size/128.xlsx')

wb1 = load_workbook(filename='../Training_Results/Final Model Training/finalResults.xlsx')

ws1 = wb1.worksheets[0]
# ws2 = wb2.worksheets[0]
# ws3 = wb3.worksheets[0]

print("Total Row in Excel 1 (excl. Header) is: ", ws1.max_row - 1)
# print("Total Row in Excel 2 (excl. Header) is: ", ws2.max_row - 1)
# print("Total Row in Excel 3 (excl. Header) is: ", ws3.max_row - 1)

# Get the number of rows (size) in 1 epoch
noOfData = -1

for col in wb1['Sheet']:
    if (col[1].value == 1):
        break;
    noOfData += 1

print("Size of 1 Epoch: ", noOfData)

incValue = 1 / noOfData
print("Incremental Value: ", incValue)

epochData = 0
epochList = []
for x in range(ws1.max_row - 1):
    epochList.append(epochData)
    epochData += incValue
print(epochList)

# Add the data to plot into a list
firstList = []
for col in wb1['Sheet']:
    firstList.append(col[4].value)
firstList.pop(0)

print(firstList)

# secondList = []
# for col in wb2['Sheet']:
#     secondList.append(col[4].value)
# secondList.pop(0)
#
# print(secondList)
#
# thirdList = []
# for col in wb3['Sheet']:
#     thirdList.append(col[4].value)
# thirdList.pop(0)
#
# print(thirdList)

# # Visualize loss history
# plt.plot(epochList, firstList, 'r-', label = "AdaDelta")
# plt.plot(epochList, secondList, 'b-',  label = "Adam")
# plt.plot(epochList, thirdList, 'g-',  label = "RMSProp")
# plt.title('Val Acc of Optimizer Experiment (10 Epochs)')

# plt.plot(epochList, firstList, 'r-', label = "0.01")
# plt.plot(epochList, secondList, 'b-',  label = "0.001")
# plt.plot(epochList, thirdList, 'g-',  label = "0.0001")
# plt.title('Train Loss of Learning Rate Experiment (10 Epochs)')

# plt.plot(epochList, firstList, 'r-', label = "32")
# plt.plot(epochList, secondList, 'b-',  label = "64")
# plt.plot(epochList, thirdList, 'g-',  label = "128")
# plt.title('Val Acc of Batch Size Experiment (10 Epochs)')

plt.plot(epochList, firstList, 'r-')
plt.axvline(x=5.72, linestyle ='--', color = 'b', label = 'Start Count \nEarly Stop Epoch:\n5.72')
plt.title('Val Acc of Final Model Experiment (20 Epochs)')


plt.legend()
plt.xlabel('Epoch')
plt.ylabel('Accuracy')
plt.margins(x=0)
# plt.margins(y=0)
# plt.xlim([0,10])
# plt.ylim([0,30])
plt.savefig("../Training_Results/Final Model Training/ValAcc.png")
plt.show();

wb1.close()
# wb2.close()
# wb3.close()
